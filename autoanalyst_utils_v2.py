"""
autoanalyst_utils_v2.py

Utilities and prompt templates for a sell-side initiation of coverage pipeline.

Design principles:
- Extract source-level research notes from each link.
- Consolidate notes into a non-redundant analytical memo.
- Generate an investment thesis and 5 key highlights first.
- Require analyst confirmation / edits on the 5 key highlights.
- Draft the remaining report sections one by one.
- Preserve the structure and discipline of a real institutional research note.
"""

from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
import trafilatura
from dotenv import load_dotenv
from openai import OpenAI
from pdfminer.high_level import extract_text
from openpyxl import load_workbook


DEFAULT_MODEL = os.getenv("AUTOANALYST_MODEL", "gpt-4o-mini")
DEFAULT_TIMEOUT = int(os.getenv("AUTOANALYST_TIMEOUT", "20"))
MAX_SOURCE_CHARS = int(os.getenv("AUTOANALYST_MAX_SOURCE_CHARS", "120000"))


SECTION_BRIEFS: Dict[str, str] = {
    "investment_thesis": (
        "This is the heart of the initiation. "
        "It should explain the core debate, the variant perception, the operating "
        "logic behind the call, and why the market should care now. "
        "The thesis must read like a seasoned sell-side analyst framing a PM-level decision."
    ),
    "key_highlights": (
        "These are the five pillars of the case. "
        "They must be mutually distinct, non-overlapping, and investment-relevant. "
        "Each highlight should be a distinct reason an institutional investor would "
        "care about the name."
    ),
    "company_overview": (
        "This is the foundational context section. "
        "It should explain what the company does, how it makes money, where it operates, "
        "and what strategic or structural features matter for the later thesis. "
        "It should be factual, but still analytical."
    ),
    "highlight_detail": (
        "This section should deepen one approved highlight and explain the mechanism, "
        "evidence, implications, and caveats. It should not re-state the thesis or "
        "the other highlights."
    ),
    "financials": (
        "This section should interpret the financial profile, not list numbers. "
        "Focus on revenue quality, growth durability, margin path, operating leverage, "
        "cash generation or burn, capital intensity, balance sheet, and what the financials "
        "imply for the equity story."
    ),
    "risks": (
        "This section should stress-test the thesis. "
        "Rank the risks by materiality and explain how each risk would transmit into the "
        "operating model, sentiment, or valuation."
    ),
}


@dataclass
class Highlight:
    title: str
    highlight: str
    implication: str = ""

    def to_dict(self) -> Dict[str, str]:
        return {
            "title": self.title,
            "highlight": self.highlight,
            "implication": self.implication,
        }


def create_openai_client(api_key: Optional[str] = None) -> OpenAI:
    """
    Create an OpenAI client using OPENAI_API_KEY unless an explicit key is provided.
    """
    load_dotenv()
    key = api_key or os.getenv("OPENAI_API_KEY")
    if not key:
        raise ValueError(
            "OPENAI_API_KEY is not set. Add it to your environment or .env file."
        )
    return OpenAI(api_key=key)


def slugify(value: str) -> str:
    """
    Convert a string into a filesystem-friendly slug.
    """
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return re.sub(r"_+", "_", value).strip("_") or "report"


def dedupe_preserve_order(items: Sequence[str]) -> List[str]:
    seen = set()
    out = []
    for item in items:
        if item not in seen:
            out.append(item)
            seen.add(item)
    return out


def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    return extract_text(BytesIO(pdf_bytes))


def _looks_like_url(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    v = value.strip().lower()
    return v.startswith("http://") or v.startswith("https://") or ".pdf" in v or "www." in v


def read_urls_from_excel(
    excel_path: str,
    sheet_name: Optional[str] = None,
    url_column: Optional[str] = None,
) -> List[str]:
    """
    Read URLs from a workbook. If url_column is omitted, the function attempts to
    infer the best column by scanning the first sheet for URL-like values.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # Explicit column path: support Excel letters only, since that is the cleanest CLI.
    if url_column:
        col_idx = None
        if re.fullmatch(r"[A-Za-z]+", url_column.strip()):
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(url_column.strip().upper())
        else:
            # Attempt header-name matching in the first row.
            header_row = 1
            for cell in ws[header_row]:
                if str(cell.value).strip().lower() == url_column.strip().lower():
                    col_idx = cell.column
                    break
        if col_idx is None:
            raise ValueError(
                f"Could not resolve url_column={url_column!r}. "
                "Use a column letter like A, B, C or a header name."
            )
        values = []
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
            if _looks_like_url(val):
                values.append(str(val).strip())
        return dedupe_preserve_order(values)

    # Infer a likely URL column by looking for URL-like density.
    best_col = None
    best_score = -1
    best_values: List[str] = []

    for col_idx in range(1, ws.max_column + 1):
        values = []
        score = 0
        for row in range(1, min(ws.max_row, 500) + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
            if _looks_like_url(val):
                score += 1
                values.append(str(val).strip())
        if score > best_score:
            best_score = score
            best_col = col_idx
            best_values = values

    if not best_values:
        # Fallback: search every cell for URLs.
        for row in ws.iter_rows():
            for cell in row:
                val = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
                if _looks_like_url(val):
                    best_values.append(str(val).strip())

    return dedupe_preserve_order(best_values)


def fetch_content_from_url(url: str, timeout: int = DEFAULT_TIMEOUT) -> str:
    """
    Fetch and extract text from a URL. Handles PDFs and HTML.
    """
    headers = {"User-Agent": "Mozilla/5.0 (compatible; AutoAnalyst/1.0)"}
    resp = requests.get(url, timeout=timeout, headers=headers)
    content_type = (resp.headers.get("Content-Type") or "").lower()

    if "application/pdf" in content_type or url.lower().endswith(".pdf"):
        return extract_text_from_pdf_bytes(resp.content)

    # HTML / web article path.
    downloaded = trafilatura.fetch_url(url)
    text = trafilatura.extract(downloaded) if downloaded else ""
    if text and text.strip():
        return text.strip()

    # Fallback to raw HTML extraction if trafilatura fetch fails.
    text = trafilatura.extract(resp.text)
    return text.strip() if text else ""


def safe_json_loads(text: str) -> Any:
    """
    Parse a JSON blob from a model response that may contain code fences.
    """
    cleaned = text.strip()

    # Remove common code fences.
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    # Try direct parse first.
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # Try to extract the first JSON object/array.
    match = re.search(r"(\{.*\}|\[.*\])", cleaned, flags=re.DOTALL)
    if match:
        return json.loads(match.group(1))

    raise ValueError("Could not parse JSON from model response.")


def _chat(
    client: OpenAI,
    system_prompt: str,
    user_prompt: str,
    model: str = DEFAULT_MODEL,
    temperature: float = 0.2,
) -> str:
    response = client.chat.completions.create(
        model=model,
        temperature=temperature,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    )
    return response.choices[0].message.content.strip()


def summarize_source(
    client: OpenAI,
    raw_text: str,
    source_label: str = "",
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Summarize one source into analytical working notes.
    """
    prompt = f"""
You are a senior sell-side equity research analyst preparing working notes for a full initiation of coverage report.

Task:
Convert the source material into analytical research notes that can be used downstream in a report. This is not a generic summary task. It is an analytical extraction task.

What matters:
- business model and operating mechanics
- revenue drivers and demand conditions
- cost structure and margin implications
- market positioning and competitive asymmetry
- capital intensity and cash generation
- balance sheet, funding, and liquidity
- strategic actions, catalysts, and execution dependencies
- risk factors and hidden constraints
- valuation relevance, if the source supports it

Rules:
- Do not paraphrase sentence-by-sentence.
- Do not write generic recap bullets.
- Every bullet must contain both a fact and an implication, unless the source is genuinely sparse.
- Prefer materiality over completeness.
- Compress repeated ideas.
- Preserve nuance and uncertainty when the source is ambiguous.
- Ignore boilerplate, marketing language, and low-signal detail.
- Do not invent information that is not in the source.

Output requirements:
- 8 to 15 bullets.
- One idea per bullet.
- Use concise institutional language.
- No intro or conclusion.
- No markdown tables.
- No numbering.

Preferred bullet format:
Observation — why it matters

SOURCE LABEL:
{source_label or "unlabeled source"}

SOURCE TEXT:
{raw_text[:MAX_SOURCE_CHARS]}
"""
    system = (
        "You write institutional-quality sell-side research notes. "
        "You are concise, analytical, and disciplined."
    )
    return _chat(client, system, prompt, model=model, temperature=0.2)


def build_knowledge_base(
    client: OpenAI,
    urls: Sequence[str],
    company_short_name: str,
    model: str = DEFAULT_MODEL,
    timeout: int = DEFAULT_TIMEOUT,
    max_sources: Optional[int] = None,
) -> str:
    """
    Ingest URLs, extract text, and build a cumulative research knowledge base.
    """
    summaries: List[str] = []
    urls = dedupe_preserve_order([u.strip() for u in urls if u and u.strip()])
    if max_sources is not None:
        urls = urls[:max_sources]

    for i, url in enumerate(urls, 1):
        try:
            print(f"[INFO] Processing source {i}/{len(urls)}: {url}")
            text = fetch_content_from_url(url, timeout=timeout)
            if not text or len(text.strip()) < 500:
                print(f"[WARN] Skipping weak or empty source: {url}")
                continue
            summary = summarize_source(
                client=client,
                raw_text=text,
                source_label=url,
                model=model,
            )
            summaries.append(f"### SOURCE {i}\nURL: {url}\n\n{summary}")
        except Exception as exc:
            print(f"[ERROR] Failed on {url}: {exc}")

    knowledge_base = "\n\n".join(summaries).strip()

    if not knowledge_base:
        raise RuntimeError("No usable sources were processed into a knowledge base.")

    return knowledge_base


def consolidate_notes(
    client: OpenAI,
    company_short_name: str,
    notes: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Merge source summaries into a single analytical memo with a clear hierarchy.
    """
    prompt = f"""
You are consolidating research notes for an institutional initiation of coverage report on {company_short_name}.

Your task is to merge overlapping observations into one master memo that can drive report writing.

This is a synthesis task, not a summarization task.

Analytical priorities:
1. What is the company really doing economically?
2. What are the core drivers of value creation or value destruction?
3. Where does the company have structural differentiation, if any?
4. What market or execution conditions matter most?
5. What are the financial implications for equity holders?
6. What is the current variant perception?
7. What are the unresolved questions that matter?

Required structure:
1. Business model and operating engine
2. Key growth drivers and catalysts
3. Competitive position and differentiation
4. Financial quality and capital efficiency
5. Balance sheet, liquidity, and funding considerations
6. Risks and constraints
7. Open questions / diligence gaps
8. Variant perception / why the market may be mispricing the story

Rules:
- Remove duplicate or near-duplicate ideas.
- Retain the strongest formulation of each idea.
- Prioritize materiality over completeness.
- Do not restate the same point under different headings.
- Avoid promotional language.
- Avoid generic industry commentary.
- Keep the memo analytical and compact.
- Use bullets under each heading.
- Do not invent facts beyond the source notes.

OUTPUT:
{notes}
"""
    system = (
        "You are a senior institutional equity research analyst. "
        "You are ruthless about materiality, logic, and non-redundancy."
    )
    return _chat(client, system, prompt, model=model, temperature=0.1)


def generate_title_and_thesis(
    client: OpenAI,
    company_short_name: str,
    company_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> Dict[str, Any]:
    """
    Generate a report title and a two-paragraph investment thesis in JSON.
    """
    prompt = f"""
You are writing the opening page of a sell-side initiation of coverage report.

Company:
{company_short_name}
Legal name:
{company_name}

Task:
Produce a concise report title and a two-paragraph investment thesis.

What the thesis must do:
- Frame the central debate in the stock.
- Explain why the company matters to an institutional investor.
- Identify the key variant perception, if one exists.
- Show the operating mechanism behind the thesis.
- Explain what the market may be underappreciating.
- Connect the story to valuation relevance or rerating potential where the evidence supports it.
- State the principal risks or constraints without turning the thesis into a risk section.
- Read like a seasoned sell-side analyst, not a marketing writer.

Tone:
- Institutional
- Analytical
- Evidence-led
- Measured
- No promotional language
- No generic superlatives

Structure:
Return valid JSON only with the following keys:
- "title": a concise report title, ideally 8-15 words
- "investment_thesis": an array of exactly 2 paragraphs
- "thesis_logic": a short sentence explaining the core logic behind the call

Constraints:
- Do not exceed two thesis paragraphs.
- Each paragraph must introduce a distinct analytical angle.
- Paragraph 1 should frame the business and central debate.
- Paragraph 2 should explain why the debate matters for valuation, catalysts, and risk-adjusted upside/downside.
- The language should be polished but not florid.

MASTER MEMO:
{memo}
"""
    system = (
        "You write Wall Street-style initiation front pages. "
        "You prioritize logic, variant perception, and investment relevance."
    )
    raw = _chat(client, system, prompt, model=model, temperature=0.2)
    data = safe_json_loads(raw)

    title = str(data.get("title", "")).strip()
    thesis = data.get("investment_thesis", [])
    if isinstance(thesis, str):
        thesis = [thesis]
    thesis = [str(x).strip() for x in thesis if str(x).strip()]

    if len(thesis) != 2:
        raise ValueError(
            "The thesis generator did not return exactly two paragraphs. "
            "Retry or inspect the raw model output."
        )

    return {
        "title": title,
        "investment_thesis": thesis,
        "thesis_logic": str(data.get("thesis_logic", "")).strip(),
        "raw": raw,
    }


def generate_key_highlights(
    client: OpenAI,
    company_short_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> Dict[str, Any]:
    """
    Generate 5 mutually distinct investment highlights in JSON.
    """
    prompt = f"""
You are drafting the five key highlights of a sell-side initiation of coverage report.

Company:
{company_short_name}

Task:
Identify the five highest-conviction pillars of the investment case.

This is the most important page of the report after the thesis. The highlights should reveal why the company deserves institutional attention, and they must not overlap.

Each highlight must:
- represent a distinct investment argument
- be material to the equity case
- be phrased like a disciplined sell-side analyst, not a marketing writer
- be specific enough to support a deeper section later in the report
- avoid duplication with the other highlights

Preferred analytical categories:
- structural growth or demand driver
- competitive differentiation or moat
- operating leverage, margin, or capital efficiency
- balance sheet, funding, liquidity, or financial resilience
- catalyst, execution advantage, or variant perception
- if relevant, a risk-adjusted asymmetry or mispricing angle

Return valid JSON only with:
- "highlights": an array of exactly 5 objects
Each object must contain:
  - "title": a concise label, 4-10 words
  - "highlight": one sentence, 20-35 words
  - "implication": one sentence explaining why it matters
  - "section_angle": a short phrase describing the unique analytical angle

Rules:
- Do not write five generic positives.
- Do not recycle the same point in different words.
- Do not use empty adjectives.
- Keep the five points non-overlapping and collectively comprehensive.
- The ordering should go from most important to least important.

MASTER MEMO:
{memo}
"""
    system = (
        "You create institutionally relevant key thesis pillars. "
        "You are precise, selective, and non-redundant."
    )
    raw = _chat(client, system, prompt, model=model, temperature=0.2)
    data = safe_json_loads(raw)
    highlights_raw = data.get("highlights", [])
    if not isinstance(highlights_raw, list):
        raise ValueError("Highlights generator did not return an array.")

    highlights: List[Highlight] = []
    for item in highlights_raw:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title", "")).strip()
        highlight = str(item.get("highlight", "")).strip()
        implication = str(item.get("implication", "")).strip()
        if title and highlight:
            highlights.append(Highlight(title=title, highlight=highlight, implication=implication))

    if len(highlights) != 5:
        raise ValueError(
            f"Expected 5 highlights, got {len(highlights)}. "
            "Retry or inspect the raw model output."
        )

    return {
        "highlights": [h.to_dict() for h in highlights],
        "raw": raw,
    }


def format_highlights_for_prompt(highlights: Sequence[Dict[str, str]]) -> str:
    lines = []
    for idx, h in enumerate(highlights, 1):
        title = h.get("title", "").strip()
        highlight = h.get("highlight", "").strip()
        implication = h.get("implication", "").strip()
        lines.append(f"{idx}. {title}\n   Highlight: {highlight}\n   Implication: {implication}")
    return "\n".join(lines)


def generate_company_overview(
    client: OpenAI,
    company_short_name: str,
    company_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    prompt = f"""
Write the Company Overview section of an initiation of coverage report on {company_short_name}.

Purpose of this section:
This is the foundation section. It should orient the reader before the thesis and highlight deep-dives. It should explain what the company does, how it makes money, where it operates, how the business is organized, and what strategic or structural features matter later in the report.

Important:
- This section is factual, but it should not be a bland recap.
- It should identify the operating logic of the business.
- It should set up the later thesis without repeating it.
- It should cover the parts of the business model that an institutional investor needs to understand to interpret the investment case.
- It should avoid generic corporate history unless it matters to the equity story.

Tone:
- Objective
- Analytical
- Institutional
- Clean prose
- No promotional language

Length target:
~700 to 900 words

Do not:
- Repeat the investment thesis
- Prematurely discuss all 5 highlights in detail
- Turn this into a risk section
- Turn this into a financial analysis section

MASTER MEMO:
{memo}
"""
    system = (
        "You write foundational company overview sections for sell-side reports. "
        "You are factual, structured, and analytical."
    )
    return _chat(client, system, prompt, model=model, temperature=0.2)


def generate_highlight_detail(
    client: OpenAI,
    company_short_name: str,
    highlight: Dict[str, str],
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    title = highlight.get("title", "").strip()
    highlight_text = highlight.get("highlight", "").strip()
    implication = highlight.get("implication", "").strip()
    angle = highlight.get("section_angle", "").strip()

    prompt = f"""
Write a deep-dive section for the following approved investment highlight in a sell-side initiation report on {company_short_name}.

Approved highlight:
Title: {title}
Highlight: {highlight_text}
Implication: {implication}
Analytical angle: {angle}

Purpose of this section:
This section must explain the highlight with depth and conviction. It should show how the evidence in the memo supports the point, why it matters economically, what it means for the stock, and what could weaken the argument.

Rules:
- Stay tightly focused on the approved highlight.
- Do not drift into the other four highlights except where necessary for context.
- Do not repeat the company overview.
- Do not repeat the thesis verbatim.
- Do not use generic filler.
- Write like a senior sell-side analyst explaining the point to an institutional investor who wants the logic, not a recap.
- Analyze cause and effect, not just facts.
- Keep the section specific, not encyclopedic.

Structure guidance:
- Open with a direct analytical judgment.
- Then support it with evidence and mechanics.
- Then explain the equity implication.
- Then note any caveat, sensitivity, or constraint.
- End without a boilerplate summary.

Length target:
~450 to 700 words

MASTER MEMO:
{memo}
"""
    system = (
        "You write tightly argued deep-dive sections for top-tier sell-side research. "
        "You maintain scope discipline and analytical depth."
    )
    return _chat(client, system, prompt, model=model, temperature=0.2)


def generate_financials_section(
    client: OpenAI,
    company_short_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    prompt = f"""
Write the Financials section for an initiation of coverage report on {company_short_name}.

Purpose of this section:
This is not a spreadsheet recap. It should interpret the financial profile and show what the numbers imply about the business quality, durability, operating leverage, and equity risk/reward.

What to cover:
- revenue quality and durability
- margin structure and margin path
- operating leverage or lack thereof
- capital intensity and reinvestment needs
- cash generation or cash burn
- liquidity, leverage, or funding need
- earnings visibility and quality of earnings
- sensitivity to growth, pricing, utilization, mix, or cost structure
- any evidence that matters for valuation or downside protection

Rules:
- Analyze, do not merely list numbers.
- If the source data is incomplete, explain the implication rather than inventing detail.
- Prioritize what matters most to an institutional investor.
- Avoid generic accounting language.
- Avoid unnecessary jargon.
- Avoid repetition with the thesis and the company overview.

Length target:
~600 to 900 words

MASTER MEMO:
{memo}
"""
    system = (
        "You write financial analysis sections for institutional research. "
        "You focus on what the financial profile says about the equity story."
    )
    return _chat(client, system, prompt, model=model, temperature=0.2)


def generate_risks_section(
    client: OpenAI,
    company_short_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    prompt = f"""
Write the Risks section for an initiation of coverage report on {company_short_name}.

Purpose of this section:
This section should stress-test the thesis and separate the real thesis-breakers from generic caveats.

Rules:
- Rank the risks by materiality.
- Distinguish structural risks, execution risks, financial risks, competitive risks, regulatory risks, and macro risks when relevant.
- Explain the mechanism by which each risk would affect the business, the sentiment, or the valuation.
- Do not write boilerplate "industry is cyclical" language unless it is material and specific.
- Do not merely restate the thesis in negative form.
- Keep the writing concise, analytical, and disciplined.

Length target:
~450 to 700 words

MASTER MEMO:
{memo}
"""
    system = (
        "You write risk sections for top-tier sell-side research. "
        "You are specific, ranked, and materially focused."
    )
    return _chat(client, system, prompt, model=model, temperature=0.2)


def final_edit_report(
    client: OpenAI,
    company_short_name: str,
    report_draft: str,
    model: str = DEFAULT_MODEL,
) -> str:
    prompt = f"""
You are the final editor of a sell-side initiation of coverage report on {company_short_name}.

Task:
Improve the report for:
- clarity
- institutional tone
- analytical sharpness
- logical flow
- section separation
- non-redundancy
- sentence precision

Rules:
- Do not add new facts.
- Do not change the thesis unless the draft is internally inconsistent.
- Remove repeated ideas across sections.
- Remove promotional language and AI-ish filler.
- Tighten phrasing where possible.
- Preserve the overall structure and voice of a real institutional research report.

Output only the edited report.

DRAFT:
{report_draft}
"""
    system = (
        "You are a meticulous sell-side editor. "
        "You improve institutional research without diluting the analysis."
    )
    return _chat(client, system, prompt, model=model, temperature=0.1)


def assemble_report_markdown(
    title: str,
    investment_thesis: Sequence[str],
    highlights: Sequence[Dict[str, str]],
    company_overview: str,
    highlight_sections: Sequence[Tuple[Dict[str, str], str]],
    financials: str,
    risks: str,
) -> str:
    """
    Assemble a clean markdown report. This is intentionally simple so the later
    Word templating phase can reuse the same structured content.
    """
    lines: List[str] = []
    lines.append(f"# {title}".strip())
    lines.append("")
    lines.append("## Investment Thesis")
    lines.append("")
    for p in investment_thesis:
        lines.append(p.strip())
        lines.append("")

    lines.append("## 5 Key Highlights")
    lines.append("")
    for idx, h in enumerate(highlights, 1):
        lines.append(f"{idx}. {h.get('title', '').strip()}")
        lines.append(f"{h.get('highlight', '').strip()}")
        if h.get("implication", "").strip():
            lines.append(f"Implication: {h.get('implication', '').strip()}")
        lines.append("")

    lines.append("## Company Overview")
    lines.append("")
    lines.append(company_overview.strip())
    lines.append("")

    for idx, (highlight, body) in enumerate(highlight_sections, 1):
        section_title = highlight.get("title", f"Highlight {idx}").strip()
        lines.append(f"## {idx}. {section_title}")
        lines.append("")
        lines.append(body.strip())
        lines.append("")

    lines.append("## Financials")
    lines.append("")
    lines.append(financials.strip())
    lines.append("")

    lines.append("## Risks")
    lines.append("")
    lines.append(risks.strip())
    lines.append("")

    return "\n".join(lines).strip() + "\n"


def save_text(path: str | Path, text: str) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def save_json(path: str | Path, data: Any) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
