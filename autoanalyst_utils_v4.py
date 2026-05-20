"""
autoanalyst_utils_v3.py

Utilities and prompt templates for a sell-side initiation of coverage pipeline.

Design principles:
- Extract source-level research notes from each link with investment-analyst discipline.
- Consolidate notes into a non-redundant, thesis-driven analytical memo.
- Generate an investment thesis and 5 key highlights — the analytical heart of the report.
- Require analyst confirmation / edits on the 5 key highlights before proceeding.
- Draft all remaining report sections with section-level context and voice.
- Produce a final clean pass that harmonizes tone, removes redundancy, and sharpens prose.
- Preserve the structure and analytical discipline of a real top-tier institutional research note.

Prompt philosophy:
- Every prompt names the analytical job, not just the content job.
- The model is asked to reason like a 25-30 year sell-side analyst, not summarize like a journalist.
- Variant perception, operating leverage, and investment relevance are required outputs.
- Generic positives are explicitly prohibited at every stage.
- Redundancy is systematically eliminated through memo structure and section scope rules.
"""

from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
import trafilatura
from dotenv import load_dotenv
from openai import OpenAI
from pdfminer.high_level import extract_text
from openpyxl import load_workbook

from bs4 import BeautifulSoup

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except Exception:
    PLAYWRIGHT_AVAILABLE = False

# ---------------------------------------------------------------------------
# ENVIRONMENT DEFAULTS
# ---------------------------------------------------------------------------

DEFAULT_MODEL = os.getenv("AUTOANALYST_MODEL", "gpt-5-nano")
DEFAULT_TIMEOUT = int(os.getenv("AUTOANALYST_TIMEOUT", "20"))
MAX_SOURCE_CHARS = int(os.getenv("AUTOANALYST_MAX_SOURCE_CHARS", "120000"))


# ---------------------------------------------------------------------------
# SECTION-LEVEL BRIEFS
# Passed into section prompts so the model understands its role in the full report.
# ---------------------------------------------------------------------------

SECTION_BRIEFS: Dict[str, str] = {
    "source_notes": (
        "The raw material of the report. Notes must convert source content into "
        "investment-relevant analytical observations: every point must carry both "
        "a fact and an implication. Pure description is waste."
    ),
    "master_memo": (
        "The synthesis layer. The memo is the internal document that drives all "
        "section writing. It must organize the best analysis from all sources into "
        "a tight, hierarchical structure with no redundancy. If the memo is weak, "
        "every section that follows it is weak."
    ),
    "investment_thesis": (
        "The most important prose in the report. A 30-year analyst reads the thesis "
        "first and decides whether the report is worth reading. It must state the "
        "central debate, identify the variant perception, explain the operating "
        "mechanism behind the call, and connect to valuation. It must not be "
        "promotional, descriptive, or encyclopedic. It must take a position."
    ),
    "key_highlights": (
        "The five pillars of the equity case. Together they must answer: why this "
        "company, why now, and what the market is missing. Each highlight must be "
        "distinct, investment-relevant, and specific enough to support a 500-word "
        "deep-dive section. They are the analytical table of contents of the report."
    ),
    "company_overview": (
        "The orientation section. It explains what the company does, how it makes "
        "money, how the business is organized, and what structural features are "
        "essential context for the thesis. It is factual-analytical, not biographical. "
        "It should set up the thesis without restating it."
    ),
    "highlight_detail": (
        "The deep-dive section for one approved highlight. Its job is to do the "
        "analytical work that makes the highlight credible: explain the mechanism, "
        "present the evidence, quantify where possible, and identify the key "
        "sensitivity or caveat. It should read like the analyst made a call and "
        "is now defending it in front of a PM."
    ),
    "financials": (
        "The financial interpretation section. Not a number recap — an interpretation "
        "of what the financial profile reveals about business quality, earnings "
        "durability, operating leverage, capital efficiency, and equity risk/reward. "
        "It should make a reader smarter about the equity story, not just the P&L."
    ),
    "risks": (
        "The stress-test section. It does not list generic industry risks. It "
        "identifies the specific mechanisms by which the thesis could be wrong, ranks "
        "them by materiality, and explains how each risk would transmit into the "
        "operating model, sentiment, or valuation. Good risk sections improve the "
        "credibility of the rest of the report."
    ),
}


# ---------------------------------------------------------------------------
# DATACLASSES
# ---------------------------------------------------------------------------

@dataclass
class Highlight:
    title: str
    highlight: str
    implication: str = ""
    section_angle: str = ""
    conviction_basis: str = ""

    def to_dict(self) -> Dict[str, str]:
        return {
            "title": self.title,
            "highlight": self.highlight,
            "implication": self.implication,
            "section_angle": self.section_angle,
            "conviction_basis": self.conviction_basis,
        }


# ---------------------------------------------------------------------------
# CLIENT + UTILITIES
# ---------------------------------------------------------------------------

def create_openai_client(api_key: Optional[str] = None) -> OpenAI:
    load_dotenv()
    key = api_key or os.getenv("OPENAI_API_KEY")
    if not key:
        raise ValueError(
            "OPENAI_API_KEY is not set. Add it to your environment or .env file."
        )
    return OpenAI(api_key=key)


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return re.sub(r"_+", "_", value).strip("_") or "report"


def dedupe_preserve_order(items: Sequence[str]) -> List[str]:
    seen = set()
    out = []
    for item in items:
        if item not in seen:
            out.append(item)
            seen.add(item)
    return out


def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    return extract_text(BytesIO(pdf_bytes))


def _looks_like_url(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    v = value.strip().lower()
    return v.startswith("http://") or v.startswith("https://") or ".pdf" in v or "www." in v


def read_urls_from_excel(
    excel_path: str,
    sheet_name: Optional[str] = None,
    url_column: Optional[str] = None,
) -> List[str]:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    if url_column:
        col_idx = None
        if re.fullmatch(r"[A-Za-z]+", url_column.strip()):
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(url_column.strip().upper())
        else:
            for cell in ws[1]:
                if str(cell.value).strip().lower() == url_column.strip().lower():
                    col_idx = cell.column
                    break
        if col_idx is None:
            raise ValueError(
                f"Could not resolve url_column={url_column!r}. "
                "Use a column letter like A, B, C or a header name."
            )
        values = []
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
            if _looks_like_url(val):
                values.append(str(val).strip())
        return dedupe_preserve_order(values)

    # Auto-detect URL column by density.
    best_col = None
    best_score = -1
    best_values: List[str] = []

    for col_idx in range(1, ws.max_column + 1):
        values = []
        score = 0
        for row in range(1, min(ws.max_row, 500) + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
            if _looks_like_url(val):
                score += 1
                values.append(str(val).strip())
        if score > best_score:
            best_score = score
            best_col = col_idx
            best_values = values

    if not best_values:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
                if _looks_like_url(val):
                    best_values.append(str(val).strip())

    return dedupe_preserve_order(best_values)


def fetch_content_from_url(url: str, timeout: int = DEFAULT_TIMEOUT) -> str:
    headers = {"User-Agent": "Mozilla/5.0 (compatible; AutoAnalyst/1.0)"}
    resp = requests.get(url, timeout=timeout, headers=headers)
    content_type = (resp.headers.get("Content-Type") or "").lower()

    if "application/pdf" in content_type or url.lower().endswith(".pdf"):
        return extract_text_from_pdf_bytes(resp.content)

    downloaded = trafilatura.fetch_url(url)
    text = trafilatura.extract(downloaded) if downloaded else ""
    if text and text.strip():
        return text.strip()

    text = trafilatura.extract(resp.text)
    return text.strip() if text else ""


# ---------------------------------------------------------------------------
# ROBUST EXTRACTION + FALLBACK INGESTION
# ---------------------------------------------------------------------------
#
# Cache layout (all under output_dir/cache/):
#
#   cache/
#     source_text/<md5>.txt      — raw scraped text for each URL
#     source_notes/<md5>.txt     — LLM analytical notes for each URL
#
# Keeping both under output_dir means:
#   • Different ticker runs never share or pollute each other's caches.
#   • The full cache travels with the output folder — easy to inspect,
#     archive, or delete per run.
#   • On restart, a URL that was fully processed (scraped + summarized)
#     costs zero API calls and zero network time.
#   • A URL that was scraped but not yet summarized re-uses the raw text
#     and skips only the network fetch.
#
# The cache root is set once per run by calling init_cache(output_dir).
# ---------------------------------------------------------------------------

_CACHE_ROOT: Path = Path("cache")   # default; overridden by init_cache()


def init_cache(output_dir: Path) -> None:
    """Call once at pipeline startup to bind the cache to the run's output_dir."""
    global _CACHE_ROOT
    _CACHE_ROOT = output_dir / "cache"
    (_CACHE_ROOT / "source_text").mkdir(parents=True, exist_ok=True)
    (_CACHE_ROOT / "source_notes").mkdir(parents=True, exist_ok=True)


def _url_digest(url: str) -> str:
    import hashlib
    return hashlib.md5(url.encode("utf-8")).hexdigest()


def _source_text_path(url: str) -> Path:
    return _CACHE_ROOT / "source_text" / f"{_url_digest(url)}.txt"


def _source_notes_path(url: str) -> Path:
    return _CACHE_ROOT / "source_notes" / f"{_url_digest(url)}.txt"


# Keep the old name as an alias so fetch_content_from_url still works.
def source_cache_path(url: str) -> Path:
    return _source_text_path(url)


def _read_cache(path: Path) -> Optional[str]:
    if path.exists():
        try:
            text = path.read_text(encoding="utf-8")
            if text.strip():
                return text
        except Exception:
            pass
    return None


def load_cached_source(url: str) -> Optional[str]:
    return _read_cache(_source_text_path(url))


def load_cached_notes(url: str) -> Optional[str]:
    """Return previously generated LLM analytical notes for this URL, or None."""
    return _read_cache(_source_notes_path(url))


def save_cached_source(url: str, text: str) -> None:
    if not text or not text.strip():
        return
    try:
        _source_text_path(url).write_text(text, encoding="utf-8")
    except Exception:
        pass


def save_cached_notes(url: str, notes: str) -> None:
    """Persist LLM analytical notes for this URL so restarts skip the API call."""
    if not notes or not notes.strip():
        return
    try:
        _source_notes_path(url).write_text(notes, encoding="utf-8")
    except Exception:
        pass


def is_weak_extraction(text: str) -> bool:

    if not text:
        return True

    text = text.strip()

    if len(text) < 1000:
        return True

    alpha_ratio = (
        sum(c.isalpha() for c in text)
        / max(len(text), 1)
    )

    if alpha_ratio < 0.50:
        return True

    lower = text.lower()

    weak_patterns = [
        "enable javascript",
        "javascript required",
        "accept cookies",
        "privacy policy",
        "terms of service",
        "loading...",
        "please wait",
        "access denied",
        "cloudflare",
        "captcha",
        "sign up",
        "subscribe",
    ]

    hits = sum(1 for p in weak_patterns if p in lower)

    if hits >= 2:
        return True

    lines = [x.strip() for x in text.splitlines() if x.strip()]

    if len(lines) < 15:
        return True

    return False


def bs4_extract(html: str) -> str:

    soup = BeautifulSoup(html, "html.parser")

    for tag in soup([
        "script",
        "style",
        "noscript",
        "header",
        "footer",
        "svg",
        "nav",
        "form",
    ]):
        tag.decompose()

    text = soup.get_text(separator="\n")

    cleaned = "\n".join(
        line.strip()
        for line in text.splitlines()
        if line.strip()
    )

    return cleaned.strip()


def playwright_extract(
    url: str,
    timeout: int = 60000,
) -> str:

    if not PLAYWRIGHT_AVAILABLE:
        raise RuntimeError(
            "Playwright is not installed. "
            "Install with: pip install playwright && playwright install"
        )

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=True
        )

        page = browser.new_page()

        page.goto(
            url,
            wait_until="networkidle",
            timeout=timeout,
        )

        try:
            page.evaluate(
                """
                () => {
                    document.querySelectorAll(
                        'script,style,noscript,header,footer,nav,svg'
                    ).forEach(el => el.remove());
                }
                """
            )
        except Exception:
            pass

        text = page.locator("body").inner_text()

        browser.close()

    return text.strip()


def extract_with_trafilatura(url: str) -> str:

    downloaded = trafilatura.fetch_url(url)

    if downloaded:
        text = trafilatura.extract(
            downloaded,
            include_comments=False,
            include_tables=True,
            no_fallback=False,
        )

        if text and text.strip():
            return text.strip()

    return ""


def extract_with_requests_bs4(
    url: str,
    timeout: int = DEFAULT_TIMEOUT,
) -> str:

    headers = {
        "User-Agent": (
            "Mozilla/5.0 "
            "(Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 "
            "(KHTML, like Gecko) "
            "Chrome/124.0 Safari/537.36"
        )
    }

    resp = requests.get(
        url,
        timeout=timeout,
        headers=headers,
    )

    resp.raise_for_status()

    return bs4_extract(resp.text)


def fetch_content_from_url(
    url: str,
    timeout: int = DEFAULT_TIMEOUT,
    use_cache: bool = True,
    use_playwright: bool = True,
) -> str:

    # -------------------------------------------------------------------
    # CACHE HIT
    # -------------------------------------------------------------------

    if use_cache:

        cached = load_cached_source(url)

        if cached and not is_weak_extraction(cached):
            return cached

    headers = {
        "User-Agent": (
            "Mozilla/5.0 "
            "(Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 "
            "(KHTML, like Gecko) "
            "Chrome/124.0 Safari/537.36"
        )
    }

    # -------------------------------------------------------------------
    # FAST PATH — PDF
    # -------------------------------------------------------------------

    try:

        resp = requests.get(
            url,
            timeout=timeout,
            headers=headers,
        )

        content_type = (
            resp.headers.get("Content-Type") or ""
        ).lower()

        if (
            "application/pdf" in content_type
            or url.lower().endswith(".pdf")
        ):

            text = extract_text_from_pdf_bytes(resp.content)

            if text and not is_weak_extraction(text):

                if use_cache:
                    save_cached_source(url, text)

                return text

    except Exception:
        pass

    # -------------------------------------------------------------------
    # FAST PATH — TRAFILATURA
    # -------------------------------------------------------------------

    try:

        downloaded = trafilatura.fetch_url(url)

        text = (
            trafilatura.extract(
                downloaded,
                include_comments=False,
                include_tables=True,
                no_fallback=False,
            )
            if downloaded
            else ""
        )

        if text and not is_weak_extraction(text):

            if use_cache:
                save_cached_source(url, text)

            return text

    except Exception:
        pass

    # -------------------------------------------------------------------
    # FAST PATH — BS4
    # -------------------------------------------------------------------

    try:

        resp = requests.get(
            url,
            timeout=timeout,
            headers=headers,
        )

        resp.raise_for_status()

        text = bs4_extract(resp.text)

        if text and not is_weak_extraction(text):

            if use_cache:
                save_cached_source(url, text)

            return text

    except Exception:
        pass

    # -------------------------------------------------------------------
    # SLOW FALLBACK — PLAYWRIGHT
    # ONLY USED IF ALL FAST PATHS FAIL
    # -------------------------------------------------------------------

    if use_playwright:

        try:

            print(
                f"[INFO] Escalating to Playwright fallback: {url}"
            )

            text = playwright_extract(url)

            if text and not is_weak_extraction(text):

                if use_cache:
                    save_cached_source(url, text)

                return text

        except Exception as exc:

            print(
                f"[WARN] Playwright fallback failed: {exc}"
            )

    return ""


def safe_json_loads(text: str) -> Any:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    match = re.search(r"(\{.*\}|\[.*\])", cleaned, flags=re.DOTALL)
    if match:
        return json.loads(match.group(1))

    raise ValueError("Could not parse JSON from model response.")


def _chat(
    client: OpenAI,
    system_prompt: str,
    user_prompt: str,
    model: str = DEFAULT_MODEL,
    temperature: float = 1,
) -> str:
    response = client.chat.completions.create(
        model=model,
        temperature=temperature,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    )
    return response.choices[0].message.content.strip()


# ---------------------------------------------------------------------------
# STEP 1 — SOURCE SUMMARIZATION
# ---------------------------------------------------------------------------

def summarize_source(
    client: OpenAI,
    raw_text: str,
    source_label: str = "",
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Convert one source into analytical working notes.

    Prompt philosophy:
    - The model is a senior analyst reading primary / secondary research, not a summarizer.
    - Every bullet must carry both a fact and an implication for the equity case.
    - The test is: could a section writer use this bullet directly?
    """
    prompt = f"""
You are a senior sell-side equity research analyst — 25 years at a top-tier institution — preparing
working notes for an initiation of coverage report. You have just read one source document.

Your task is to extract the highest-value analytical content from this source and convert it into
crisp, investment-relevant research notes that a report writer can use directly.

This is NOT a summarization task. It is an analytical extraction task.

THE TEST FOR EVERY BULLET:
Ask yourself: "If I included this in a sell-side report, would a buy-side PM learn something
that changes how they think about this equity?" If the answer is no, cut it.

WHAT TO LOOK FOR AND EXTRACT:
- Operating model mechanics: how does the business actually make money, and what drives unit economics?
- Revenue quality and durability: contract length, renewal dynamics, mix shift, pricing power
- Cost structure: labor intensity, fixed vs. variable, cost-to-serve, operating leverage
- Competitive positioning: structural advantages, moats, barriers to replication, market share data
- Capital efficiency: asset turns, capex intensity, returns on invested capital
- Balance sheet and liquidity: debt structure, covenants, runway, funding needs
- Earnings quality: cash conversion, non-cash items, one-time adjustments, reserve policies
- Strategic initiatives: acquisitions, new verticals, tech investments — with a view on whether they create or destroy value
- Catalysts: near-term events, contract wins, regulatory changes, industry inflections
- Risks and constraints: what could break the thesis, and by how much?
- Variant perception clues: what is the market likely pricing in vs. what the source suggests?

RULES:
- Do not paraphrase sentence-by-sentence. Synthesize and re-express analytically.
- Never write a bullet that is pure description with no implication.
- Every bullet: [observation] — [why it matters for the equity or the analytical case]
- Prioritize specificity: named metrics, named competitors, named contracts beat generic commentary.
- Compress near-duplicate ideas into one superior bullet.
- Ignore all boilerplate, marketing language, forward-looking safe harbors, and zero-signal text.
- Do not invent facts not in the source. Flag uncertainty where it exists.
- If the source is thin or low-quality, produce fewer bullets and flag it.

OUTPUT REQUIREMENTS:
- 8 to 15 bullets
- One material idea per bullet
- Institutional register: no casual language, no superlatives, no adjective inflation
- No intro line, no conclusion line, no markdown table
- No bullet numbering

BULLET FORMAT:
[Crisp factual observation] — [investment implication or analytical significance]

SOURCE LABEL:
{source_label or "unlabeled source"}

SOURCE TEXT:
{raw_text[:MAX_SOURCE_CHARS]}
"""
    system = (
        "You are a managing director-level sell-side equity research analyst. "
        "You have spent 25 years distilling information into investment-grade insight. "
        "You are precise, selective, and allergic to generic commentary."
    )
    return _chat(client, system, prompt, model=model, temperature=1)


# ---------------------------------------------------------------------------
# STEP 2 — KNOWLEDGE BASE ASSEMBLY
# ---------------------------------------------------------------------------

def build_knowledge_base(
    client: OpenAI,
    urls: Sequence[str],
    company_short_name: str,
    model: str = DEFAULT_MODEL,
    timeout: int = DEFAULT_TIMEOUT,
    max_sources: Optional[int] = None,
    use_playwright: bool = True,
    output_dir: Optional[Path] = None,
) -> str:
    # Bind the cache to this run's output directory so different ticker runs
    # never share caches, and a full restart can skip already-done URLs.
    if output_dir is not None:
        init_cache(output_dir)

    summaries: List[str] = []

    urls = dedupe_preserve_order([
        u.strip()
        for u in urls
        if u and u.strip()
    ])

    if max_sources is not None:
        urls = urls[:max_sources]

    processed = 0
    from_cache = 0
    skipped = 0
    failed = 0

    for i, url in enumerate(urls, 1):

        try:
            print(f"[INFO] Source {i}/{len(urls)}: {url}")

            # ------------------------------------------------------------------
            # NOTES CACHE HIT — both scrape and summarization already done
            # ------------------------------------------------------------------
            cached_notes = load_cached_notes(url)
            if cached_notes:
                print(f"[CACHE] Notes cache hit — skipping scrape + summarization")
                summaries.append(
                    f"### SOURCE {i}\n"
                    f"URL: {url}\n\n"
                    f"{cached_notes}"
                )
                from_cache += 1
                continue

            # ------------------------------------------------------------------
            # SCRAPE (uses its own text cache internally)
            # ------------------------------------------------------------------
            text = fetch_content_from_url(
                url=url,
                timeout=timeout,
                use_cache=True,
                use_playwright=use_playwright,
            )

            if is_weak_extraction(text):
                skipped += 1
                print(f"[WARN] Weak extraction after all fallbacks: {url}")
                continue

            print(f"[INFO] Extracted {len(text):,} chars")

            # ------------------------------------------------------------------
            # SUMMARIZE — LLM call; result is cached immediately after
            # ------------------------------------------------------------------
            summary = summarize_source(
                client=client,
                raw_text=text,
                source_label=url,
                model=model,
            )

            save_cached_notes(url, summary)

            summaries.append(
                f"### SOURCE {i}\n"
                f"URL: {url}\n\n"
                f"{summary}"
            )

            processed += 1

        except Exception as exc:
            failed += 1
            print(f"[ERROR] Failed on {url}: {exc}")

    knowledge_base = "\n\n".join(summaries).strip()

    print("\n" + "=" * 80)
    print("[INFO] KNOWLEDGE BASE BUILD SUMMARY")
    print("=" * 80)
    print(f"Processed (new)        : {processed}")
    print(f"Loaded from cache      : {from_cache}")
    print(f"Weak / Skipped         : {skipped}")
    print(f"Failed                 : {failed}")
    print(f"Total URLs             : {len(urls)}")
    print("=" * 80 + "\n")

    if not knowledge_base:
        raise RuntimeError(
            "No usable sources were processed into a knowledge base."
        )

    return knowledge_base


# ---------------------------------------------------------------------------
# STEP 3 — MASTER MEMO CONSOLIDATION
# ---------------------------------------------------------------------------

def consolidate_notes(
    client: OpenAI,
    company_short_name: str,
    notes: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Merge all source-level notes into one master analytical memo.

    Prompt philosophy:
    - The memo is the single source of truth for all section writing.
    - Every section prompt receives the memo as its primary input.
    - The memo must be structured, non-redundant, and thesis-driving.
    - Weak memos produce weak reports. This step is not cosmetic.
    """
    prompt = f"""
You are a senior equity research analyst consolidating all source-level notes into a single master
analytical memo that will drive every section of an initiation of coverage report on {company_short_name}.

This memo is your internal research document. It is not a section of the report — it is the engine
that powers the report. Every section writer will use this memo, so it must be organized, complete,
and ruthlessly non-redundant.

YOUR JOB IS SYNTHESIS, NOT SUMMARIZATION.
Summarization = listing what each source said.
Synthesis = identifying what the totality of evidence implies about the business, the equity, and the variant perception.

ANALYTICAL QUESTIONS THAT MUST BE ANSWERED IN THE MEMO:
1. What is this company's actual economic engine — how does it generate value, for whom, and how defensibly?
2. What are the primary drivers of revenue growth — are they structural/secular or cyclical/executional?
3. Where does the company have genuine competitive differentiation vs. where is it just scale?
4. What does the financial profile reveal about business quality, capital efficiency, and earnings durability?
5. What is the balance sheet situation — overfunded, constrained, opportunistic?
6. What are the principal risks — and are they thesis-breakers or manageable sensitivities?
7. What unresolved questions would a skeptical PM ask that the sources cannot yet answer?
8. What is the variant perception — where does consensus appear to be wrong, and why?

REQUIRED MEMO STRUCTURE (8 sections, each with analytical bullet points):

1. Business Model and Operating Engine
   - How the company makes money, unit economics, revenue model, contract structure
   - What drives margin: fixed vs. variable cost structure, self-perform vs. subcontracted, pricing levers
   - How the model scales and what happens to unit economics as it does

2. Growth Drivers and Near-Term Catalysts
   - Structural demand drivers (secular, regulatory, technology-driven)
   - Company-specific catalysts: contract wins, new verticals, M&A contributions, product launches
   - Management's own guidance and the assumptions embedded in it

3. Competitive Position and Moat Analysis
   - Where the company wins vs. competitors and why
   - Barriers to replication: scale, technology, regulatory, switching cost, network effect
   - Where competitive advantage is real vs. claimed — be skeptical

4. Financial Quality and Capital Efficiency
   - Revenue quality and visibility (contract duration, renewal rates, ARR-like characteristics)
   - Margin trajectory: where margins are going and what gets them there
   - Capital efficiency: ROIC, asset intensity, reinvestment rate, operating leverage
   - Earnings quality: cash conversion, adjustments, non-recurring items

5. Balance Sheet, Liquidity, and Funding
   - Debt structure: maturity profile, covenants, leverage ratios
   - Liquidity position: cash, revolver availability, near-term obligations
   - Capital allocation priorities: M&A, dividends, buybacks, organic capex
   - Financial flexibility vs. financial constraint

6. Risks and Constraints
   - Ranked by materiality and thesis-relevance
   - For each risk: mechanism (how it would hurt), magnitude (how much), and mitigant (if any)
   - Structural vs. cyclical vs. execution vs. regulatory risks — label them

7. Open Questions and Diligence Gaps
   - What the sources could not answer that a PM would demand
   - Where the evidence is thin, conflicting, or potentially out of date
   - What management commentary is needed to fill the gaps

8. Variant Perception / What the Market May Be Missing
   - Where does the current market valuation appear to under-appreciate or over-price an element?
   - Is there a re-rating catalyst, a sentiment inflection, or a consensus misconception?
   - What would a bull say that the bear doesn't credit — and is it evidence-supported?

RULES:
- Remove all near-duplicate ideas. Keep the best version of each insight.
- Never restate the same observation under two different headings.
- Prioritize materiality: a weak insight that is covered by every analyst adds no value.
- Flag genuine uncertainty where sources conflict or are ambiguous. Do not paper over it.
- Do not invent facts. Every claim must trace to the source notes.
- Use precise, institutional language throughout.
- Use bullet points under each heading — 4 to 10 bullets per section.
- Do not write an intro or conclusion to the memo.

SOURCE NOTES TO SYNTHESIZE:
{notes}
"""
    system = (
        "You are a managing director at a top-tier sell-side research firm. "
        "You are synthesizing all available research into a definitive analytical memo "
        "that will guide an initiation of coverage report. You are ruthless about "
        "materiality, logic, and non-redundancy. Weak memos produce weak reports."
    )
    return _chat(client, system, prompt, model=model, temperature=1)


# ---------------------------------------------------------------------------
# STEP 4A — INVESTMENT THESIS + TITLE
# ---------------------------------------------------------------------------

def generate_title_and_thesis(
    client: OpenAI,
    company_short_name: str,
    company_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> Dict[str, Any]:
    """
    Generate the report title and investment thesis.

    Prompt philosophy:
    - The thesis is the most-read and most-judged element of the report.
    - A PM reads the thesis and decides in 30 seconds whether to read the rest.
    - It must name the debate, state the variant perception, explain the mechanism,
      and connect to a valuation conclusion or catalyst setup — all without being promotional.
    - The title must function as the thesis summary in 10-14 words.
    """
    prompt = f"""
You are opening an institutional sell-side initiation of coverage report on {company_name} ({company_short_name}).

You are drafting the report title and the investment thesis — the two most important elements
of the report. A portfolio manager will read these first and decide within 30 seconds whether
the report deserves their time. If the thesis is generic, descriptive, or promotional, it fails.

WHAT A GREAT THESIS DOES:
1. Names the central debate: what is the market argument you are joining, and where do you land?
2. Identifies the variant perception: what do you believe that consensus is getting wrong — and why?
3. Explains the operating mechanism: what is the business doing that creates the investment opportunity?
4. Connects to valuation or catalysts: why does the debate resolve itself in the direction of your call?
5. Acknowledges the principal constraint: what could make you wrong — without turning into a risk section?
6. Reads like a seasoned analyst talking to a PM, not a marketing writer talking to a retail investor.

WHAT KILLS A THESIS:
- "The company is a leading provider of X in a large and growing market." (description, not analysis)
- "We initiate with a Buy rating as we see significant upside potential." (circular, zero content)
- "Strong management team with decades of experience." (filler)
- Any sentence that would be equally true for 50 other companies in the sector.
- Promotional adjectives: transformative, game-changing, best-in-class, world-class, unique.
- Restating the 5 highlights in abbreviated form (the thesis is upstream of the highlights).

THESIS STRUCTURE:
- Paragraph 1: Frame the business and the central debate. What is this company doing economically?
  What is the market's current assumption, and where does it appear mis-calibrated?
  Show the operating mechanism that underpins the call.
- Paragraph 2: Explain why the debate matters for the equity now. What catalysts, inflection points,
  or valuation dynamics make this a timely call? What risk-adjusted upside/downside does the setup offer?
  Close with what would make you wrong — one sentence, honest, not defensive.

TITLE REQUIREMENTS:
- 8 to 14 words
- Should encode the investment thesis in compressed form — a reader should understand the call
  from the title alone
- Avoid generic sector descriptors ("leading", "dominant", "premier")
- Factual, not promotional; analytical, not cheerleading
- Examples of strong report titles (not for copying):
  * "Structural Pricing Power Meets a Cyclical Discount: The Re-Rating Case for [Co]"
  * "Data Advantage Compounding Quietly: [Co]'s Path from Scale to Margin"
  * "The Market Is Paying for Commodity Revenue; The Business Is Building a Moat"

TONE:
- Institutional and measured
- Analytical confidence without promotional certainty
- Evidence-led, not adjective-led
- The analyst has a view and is defending it, not selling it

OUTPUT FORMAT:
Return valid JSON only with these keys:
- "title": string (8-14 words)
- "investment_thesis": array of exactly 2 paragraph strings
- "thesis_logic": one sentence capturing the core analytical logic of the call (used internally)
- "variant_perception": one sentence stating the specific consensus mispricing or misunderstanding

CONSTRAINT:
- Do not exceed 2 thesis paragraphs.
- Each paragraph must introduce a distinct analytical angle.
- Do not use the company's own marketing language.
- Do not reference the 5 highlights by name or number.

MASTER MEMO:
{memo}
"""
    system = (
        "You are a managing director and senior research analyst at a top-tier sell-side firm. "
        "You have opened more initiation reports than you can count and you know that the thesis "
        "is the only thing that matters for first impressions. You write with conviction, precision, "
        "and the analytical authority that comes from having read every number twice."
    )
    raw = _chat(client, system, prompt, model=model, temperature=1)
    data = safe_json_loads(raw)

    title = str(data.get("title", "")).strip()
    thesis = data.get("investment_thesis", [])
    if isinstance(thesis, str):
        thesis = [thesis]
    thesis = [str(x).strip() for x in thesis if str(x).strip()]

    if len(thesis) != 2:
        raise ValueError(
            "The thesis generator did not return exactly two paragraphs. "
            "Retry or inspect the raw model output."
        )

    return {
        "title": title,
        "investment_thesis": thesis,
        "thesis_logic": str(data.get("thesis_logic", "")).strip(),
        "variant_perception": str(data.get("variant_perception", "")).strip(),
        "raw": raw,
    }


# ---------------------------------------------------------------------------
# STEP 4B — 5 KEY HIGHLIGHTS
# ---------------------------------------------------------------------------

def generate_key_highlights(
    client: OpenAI,
    company_short_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> Dict[str, Any]:
    """
    Generate the 5 key investment highlights.

    Prompt philosophy:
    - The 5 highlights are the analytical table of contents of the report.
    - Each highlight must answer a different question a PM would ask.
    - Together they must be collectively exhaustive without being individually redundant.
    - A weak set of highlights signals a weak analyst. A strong set signals conviction and depth.
    - The ordering matters: most important to least important, not easiest to hardest.
    """
    prompt = f"""
You are a senior sell-side research analyst drafting the five key highlights for an initiation of
coverage report on {company_short_name}.

These are the five most important reasons an institutional investor should pay attention to this name.
They appear immediately after the investment thesis and serve as the analytical table of contents
for the entire report. Each highlight will become a 500-700 word deep-dive section.

THE STANDARD EVERY HIGHLIGHT MUST MEET:
A buy-side PM reads a highlight and says: "That's a specific, investable observation I didn't
fully appreciate. I want to read the deep-dive section." If instead they say "that's obvious"
or "that's vague", the highlight has failed.

WHAT MAKES A GREAT HIGHLIGHT:
- It is investment-specific, not sector-generic. It applies to this company in a way that does
  not apply equally to its competitors.
- It identifies a mechanism, not just a trait. Not "strong margin profile" but "self-perform model
  eliminates third-party markup and converts labor scale into operating leverage."
- It has a clear equity implication. The reader understands why it matters for the stock.
- It is phrased with analytical precision. No superlatives, no marketing language.
- It is ordered — highlight 1 is the most important reason to own the stock.

WHAT KILLS A HIGHLIGHT:
- "Large addressable market with secular tailwinds." (generic for every growth company)
- "Experienced management team." (zero analytical content)
- "Diversified revenue base reduces risk." (true of every conglomerate, says nothing specific)
- Two highlights that are really one highlight split in two.
- A highlight that repeats a point made in the investment thesis at the same level of depth.

ANALYTICAL CATEGORIES TO CONSIDER (pick the five most relevant to this company):
- Structural demand driver: a secular shift in end-market demand that the company is positioned to capture
- Competitive moat or differentiation: a specific mechanism that prevents share loss or enables premium pricing
- Operating leverage or margin expansion: a clear path from current margins to a structurally higher level
- Financial quality and capital efficiency: cash generation, ROIC, balance sheet strength, or capital return
- Execution catalyst or re-rating driver: a specific event, contract, or strategic move that unlocks value
- Mispricing or variant perception: a specific way the current price appears to embed wrong assumptions
- Technology or platform advantage: a proprietary capability that creates compounding returns
- M&A or capital deployment optionality: a well-funded balance sheet targeting value-accretive deployment

INSTRUCTIONS FOR ORDERING:
- Place the highlight that is most directly connected to the investment thesis first.
- Place the risk-adjacent or caveat-adjacent highlight last.
- The order should feel like a logical argument building from foundation to conclusion.

OUTPUT FORMAT:
Return valid JSON only with:
- "highlights": array of exactly 5 objects
Each object:
  - "title": 5-10 words, analytical and specific (no generic descriptors)
  - "highlight": 20-40 words, one tight analytical sentence with a fact and an implication
  - "implication": 15-30 words, one sentence on why this moves the equity case
  - "section_angle": 10-20 words describing the unique analytical angle for the deep-dive section
  - "conviction_basis": one phrase identifying the primary evidence base for this highlight

ANTI-REDUNDANCY CHECK (before you finalize):
- Read all 5 highlights together. Does any highlight say essentially the same thing as another?
  If yes, collapse them and replace the redundant one with a distinct argument.
- Do all 5 highlights reference the same segment or mechanism? If yes, diversify.
- Does the set collectively explain the full investment case? If important elements of the memo
  are unaddressed, revise.

MASTER MEMO:
{memo}
"""
    system = (
        "You are a managing director-level sell-side analyst who has built hundreds of investment cases. "
        "You know that the 5 key highlights are the most analytically demanding page of an initiation report. "
        "You write highlights that are specific, non-overlapping, investment-relevant, and ordered "
        "by conviction. You never write a generic positive about a company."
    )
    raw = _chat(client, system, prompt, model=model, temperature=1)
    data = safe_json_loads(raw)
    highlights_raw = data.get("highlights", [])
    if not isinstance(highlights_raw, list):
        raise ValueError("Highlights generator did not return an array.")

    highlights: List[Highlight] = []
    for item in highlights_raw:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title", "")).strip()
        highlight_text = str(item.get("highlight", "")).strip()
        implication = str(item.get("implication", "")).strip()
        section_angle = str(item.get("section_angle", "")).strip()
        conviction_basis = str(item.get("conviction_basis", "")).strip()
        if title and highlight_text:
            highlights.append(Highlight(
                title=title,
                highlight=highlight_text,
                implication=implication,
                section_angle=section_angle,
                conviction_basis=conviction_basis,
            ))

    if len(highlights) != 5:
        raise ValueError(
            f"Expected 5 highlights, got {len(highlights)}. "
            "Retry or inspect the raw model output."
        )

    return {
        "highlights": [h.to_dict() for h in highlights],
        "raw": raw,
    }


# ---------------------------------------------------------------------------
# STEP 5A — COMPANY OVERVIEW
# ---------------------------------------------------------------------------

def generate_company_overview(
    client: OpenAI,
    company_short_name: str,
    company_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Write the Company Overview section.

    Role in the report:
    - Comes AFTER the thesis and highlights in the full report sequence.
    - Its job is to give the reader the business context they need to understand
      the thesis and highlights at a deeper level.
    - Factual-analytical: it explains the business model, revenue streams, segment
      structure, and strategic priorities — without editorializing prematurely.
    - It is NOT a biography. Corporate history matters only if it explains the
      current business model or competitive position.

    Prompt philosophy:
    - The reader already knows the thesis. The overview makes it operational.
    - Segment structure, revenue mix, customer concentration, and operating model
      are the key deliverables.
    - The section should close by setting up the themes the five highlights will explore.
    """
    prompt = f"""
You are writing the Company Overview section of an initiation of coverage report on {company_name} ({company_short_name}).

SECTION ROLE IN THE FULL REPORT:
This section appears after the investment thesis and 5 key highlights. The reader already knows
your call. Now they need the business context to understand how the thesis works mechanically.
This is an orientation section — thorough, analytical, and factual — but it does not argue.
The thesis argues. The highlights argue. The overview explains.

WHAT THIS SECTION MUST COVER:
1. Business description: What does the company do? In plain analytical language, explain the
   core business activity, the customers it serves, and the economic model.
2. Revenue model: How does the company get paid? Contract-based, project-based, subscription,
   volume-tiered, milestone? What determines revenue velocity?
3. Segment structure: How is the business organized? What does each segment contribute in
   revenue and profit? Which segments are growing vs. mature vs. declining?
4. Operating model: Self-perform vs. asset-light? Labor-intensive vs. capital-intensive?
   What are the key inputs and how does the company manage them?
5. Customer profile: Who are the clients? How concentrated is revenue? What is the retention
   or renewal dynamic?
6. Geographic footprint and market position: Where does the company operate and what is its
   competitive standing in each geography or vertical?
7. Strategic direction: What are the stated strategic priorities and what do they tell an analyst
   about where management is allocating capital and attention?

WHAT THIS SECTION MUST NOT DO:
- Argue the investment case (that is the thesis's job).
- Pre-empt the 5 highlight deep-dives (those sections expand each pillar independently).
- Discuss risks in detail (risks get their own section).
- Provide a detailed financial analysis (that is the financials section's job).
- Recount corporate history unless it directly explains the current operating model.
- Use promotional language borrowed from the company's own marketing materials.

TONE AND STYLE:
- Factual and objective, but not bland.
- The analytical instinct should come through in how you organize and prioritize information —
  lead with what matters most for the equity story, not with what came first chronologically.
- Clean, institutional prose. No bullet points. No subheadings.
- A senior analyst at a competing firm should read this and say: "That's a fair, complete picture
  of how this business works."

LENGTH TARGET: 700 to 900 words

MASTER MEMO:
{memo}
"""
    system = (
        "You write company overview sections for institutional equity research reports. "
        "You are factual, analytically organized, and efficient with language. "
        "You never confuse the overview's job (explain the business) with the thesis's "
        "job (argue the investment case)."
    )
    return _chat(client, system, prompt, model=model, temperature=1)


# ---------------------------------------------------------------------------
# STEP 5B — HIGHLIGHT DEEP-DIVES
# ---------------------------------------------------------------------------

def generate_highlight_detail(
    client: OpenAI,
    company_short_name: str,
    highlight: Dict[str, str],
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Write a deep-dive section for one approved highlight.

    Role in the report:
    - This is where the analyst earns their credibility.
    - The highlight stated a conviction. This section must prove it.
    - Mechanism + evidence + implication + caveat is the structure.
    - The reader should finish this section with a clearer, more confident
      understanding of why the highlight matters — or a clear view of what
      would need to be true for it to matter even more / matter less.

    Prompt philosophy:
    - The section is not a summary of the highlight — it is the supporting argument.
    - Quantification is strongly preferred over assertion.
    - The caveat or sensitivity is required — it improves credibility.
    - The section must not re-argue the thesis or bleed into other highlights.
    """
    title = highlight.get("title", "").strip()
    highlight_text = highlight.get("highlight", "").strip()
    implication = highlight.get("implication", "").strip()
    angle = highlight.get("section_angle", "").strip()
    conviction_basis = highlight.get("conviction_basis", "").strip()

    prompt = f"""
You are writing a deep-dive section for the following approved investment highlight in an
initiation of coverage report on {company_short_name}.

APPROVED HIGHLIGHT:
Title: {title}
Highlight sentence: {highlight_text}
Implication: {implication}
Analytical angle for this section: {angle}
Primary evidence base: {conviction_basis}

SECTION ROLE IN THE FULL REPORT:
This section appears after the Company Overview and alongside the four other highlight sections.
The reader has already seen the highlight summary on the thesis page. This section is where
the highlight is proven, not re-stated. The reader expects depth, evidence, and conviction.
They do not expect a repetition of what they already read.

THE ANALYTICAL STRUCTURE YOU MUST FOLLOW:

1. OPENING JUDGMENT (1 paragraph):
   Open with a direct, declarative analytical statement — your strongest version of the
   investment argument for this highlight. Do not open with a company description. Do not
   open with a vague setup. Open with the claim. Example approach:
   "[Company]'s [specific mechanism] represents a [structural / durable / underappreciated]
   advantage that we believe the current consensus does not fully credit."

2. MECHANISM AND EVIDENCE (2-3 paragraphs):
   Explain HOW the highlight works mechanically. What are the inputs, outputs, and key
   variables? Then present the best evidence from the memo: specific data points, named
   contracts, named competitors, financial metrics, operating statistics. Quantify where
   possible. Prefer specifics ("$125M contract at MIA airport, 5-year term") over generics
   ("recent large contract win"). Explain cause and effect, not just correlation.

3. EQUITY IMPLICATION (1 paragraph):
   Now connect the mechanism and evidence to the equity case. What does this mean for
   revenue trajectory, margin profile, ROIC, valuation, or re-rating potential? Be specific
   about the channel of impact. Avoid saying "this is positive for the stock" — say why and
   by what mechanism.

4. CAVEAT OR SENSITIVITY (1 paragraph):
   Every honest analyst identifies the condition under which their argument is weaker.
   What would need to be true for this highlight to matter less than you expect?
   What is the key sensitivity — execution risk, macro dependency, competitive response?
   Note it once, clearly. This is not a risk section — one focused caveat improves
   the credibility of the overall argument.

CRITICAL SCOPE RULES:
- Stay within the scope of this highlight. Do not drift into the other four highlights.
- Do not repeat the company overview. Assume the reader has read it.
- Do not re-state the investment thesis at length.
- Do not use the highlight title as your opening line.
- Do not end with a boilerplate summary or a call to action.

STYLE:
- Write in the first-person institutional plural (we believe, we see, we estimate).
- Analytical confidence backed by evidence — not assertion alone.
- No bullet points within the section. Clean prose paragraphs only.
- No subheadings within the section.
- Every paragraph must advance the argument — no filler paragraphs.

LENGTH TARGET: 500 to 700 words

MASTER MEMO:
{memo}
"""
    system = (
        "You are a senior sell-side analyst writing a deep-dive section for a specific investment "
        "highlight. Your job is to prove the highlight, not re-state it. You write with analytical "
        "conviction, evidence discipline, and scope precision. You end every section slightly earlier "
        "than you think you should — tight, not exhaustive."
    )
    return _chat(client, system, prompt, model=model, temperature=1)


# ---------------------------------------------------------------------------
# STEP 5C — FINANCIALS
# ---------------------------------------------------------------------------

def generate_financials_section(
    client: OpenAI,
    company_short_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Write the Financials section.

    Role in the report:
    - Comes after the 5 highlight deep-dives.
    - Its job is to interpret the financial profile, not recap the income statement.
    - The test: after reading this, does the investor understand whether the financial
      model supports or constrains the thesis?

    Prompt philosophy:
    - Financial analysis, not financial description.
    - Revenue quality, margin path, operating leverage, and cash generation are the
      core analytical deliverables.
    - Valuation color belongs here — even without a formal price target, the section
      should help the reader think about how to frame the risk/reward.
    """
    prompt = f"""
You are writing the Financials section of an initiation of coverage report on {company_short_name}.

SECTION ROLE IN THE FULL REPORT:
This section appears after the five highlight deep-dives. The reader has already absorbed
the investment thesis and the five analytical pillars. The financials section is where you
show that the numbers support — or at least do not contradict — the thesis. It is an
interpretive section, not a data-delivery section.

THE CRITICAL DISTINCTION:
A bad financials section lists numbers. A good financials section interprets what the numbers
reveal about the quality, durability, and trajectory of the business, and connects the
financial profile to the equity story.

WHAT TO ANALYZE AND WHAT TO SAY ABOUT IT:

1. Revenue Quality and Visibility:
   - What proportion of revenue is contractual, recurring, or quasi-recurring?
   - What is the revenue growth rate and what drives it — volume, price, mix, or M&A?
   - How visible is the revenue trajectory? Are there leading indicators (backlog, renewal rate, pipeline)?
   - Is revenue quality improving or degrading? Why?

2. Margin Structure and Path:
   - What are the current EBITDA, EBIT, and net margins and how do they compare to peers?
   - What are the primary drivers of margin — labor as % of revenue, overhead absorption, pricing?
   - Is there a credible path to margin expansion — operating leverage, mix shift, tech-enabled efficiency?
   - What are the main risks to the margin trajectory?

3. Operating Leverage:
   - How does the business respond to incremental revenue? Does margin expand meaningfully?
   - What is the fixed cost base and how does it behave across cycles?
   - Is there evidence of operating leverage in the historical record?

4. Capital Efficiency and ROIC:
   - What is the capital intensity of the business — high or low? Asset-heavy or asset-light?
   - What is the return on invested capital and how does it compare to the cost of capital?
   - How does management deploy incremental capital — organically or through M&A?
   - What is the reinvestment rate and what does it imply for long-run growth?

5. Cash Generation and Conversion:
   - What is the free cash flow profile — generation rate, variability, conversion from net income?
   - What are the main uses of FCF — capex, M&A, dividends, buybacks?
   - Is FCF growing, stable, or pressured? Why?

6. Balance Sheet and Financial Resilience:
   - What is the net leverage ratio and how does it compare to sector norms?
   - What is the debt maturity profile and are there near-term refinancing risks?
   - Does the company have adequate liquidity for its current strategic agenda?
   - Is the balance sheet a competitive asset (M&A optionality) or a constraint?

7. Earnings Quality and Investor Communication:
   - Are there significant gaps between GAAP earnings and adjusted earnings? Why?
   - How does the company communicate financial performance — are there credibility concerns?
   - Are management's stated targets realistic based on the available financial data?

WHAT TO AVOID:
- Do not list raw numbers without interpreting them.
- Do not write "revenue grew X% to $Y billion" and stop there — explain what drove the growth
  and what it implies.
- Do not ignore gaps, inconsistencies, or unflattering data points in the memo.
- Do not repeat analysis already covered in the highlight sections.
- Do not use accounting jargon without explaining what it implies for the equity investor.

STYLE:
- Clean institutional prose. No bullet points. No subheadings within the body.
- Write in the first-person institutional plural (we note, we believe, we estimate).
- Analytical and precise: favor specific numbers and named drivers over vague characterizations.
- Intellectually honest: acknowledge weaknesses in the financial profile where they exist.

LENGTH TARGET: 650 to 950 words

MASTER MEMO:
{memo}
"""
    system = (
        "You write financial analysis sections for institutional equity research. "
        "You interpret financial profiles — you do not list numbers. "
        "Every financial observation must connect to the equity story: "
        "what does this metric tell us about the quality, trajectory, or risk of the business?"
    )
    return _chat(client, system, prompt, model=model, temperature=1)


# ---------------------------------------------------------------------------
# STEP 5D — RISKS
# ---------------------------------------------------------------------------

def generate_risks_section(
    client: OpenAI,
    company_short_name: str,
    memo: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Write the Risks section.

    Role in the report:
    - Comes last in the analytical body of the report.
    - Its purpose is to stress-test the thesis and identify the conditions under
      which the investment case would be wrong, and by how much.
    - A well-written risks section increases the credibility of the entire report.
    - A boilerplate risks section signals that the analyst did not think hard enough.

    Prompt philosophy:
    - Risks must be ranked by materiality (thesis-impact), not by type.
    - Every risk must have a transmission mechanism: HOW it hurts, not just THAT it could hurt.
    - Generic industry risks are only acceptable if they are quantifiably material to this company.
    - The risks section is not a disclaimer — it is a stress test.
    """
    prompt = f"""
You are writing the Risks section of an initiation of coverage report on {company_short_name}.

SECTION ROLE IN THE FULL REPORT:
This is the final analytical section of the report. The reader has absorbed the thesis, the
five highlights, the company overview, and the financial analysis. They believe the case.
Now they need to know what could make the case wrong. This section tells them.

THE STANDARD FOR A GOOD RISKS SECTION:
A great risks section does not protect the analyst legally. It protects the investor analytically.
It identifies the specific conditions, mechanisms, and magnitudes that would undermine the thesis —
and it is honest about which risks are manageable sensitivities vs. genuine thesis-breakers.

A poor risks section lists "competition", "regulation", and "macroeconomic conditions" with
no specifics and no transmission mechanisms. Do not write that section.

HOW TO STRUCTURE AND RANK THE RISKS:

RANKING PRINCIPLE: Order risks by how much they would damage the thesis, not by how likely
they are. A low-probability thesis-killer ranks above a high-probability noise event.

FOR EACH RISK, INCLUDE:
1. Risk label: a specific, descriptive title (not just "labor risk" but "structural labor cost
   inflation driven by union bargaining and minimum-wage legislation")
2. Transmission mechanism: HOW does this risk hurt — revenue? margin? cash flow? valuation multiple?
3. Materiality: approximately how much could it move the financial profile if it crystallizes?
4. Mitigant: is there a natural hedge, contractual protection, or management action that limits
   the damage? Be honest about whether the mitigant is adequate.

RISK TAXONOMY (use what is relevant, do not force all categories):
- Structural risks: permanently change the business model or competitive position
- Execution risks: management fails to deliver on stated strategic priorities
- Financial risks: balance sheet, liquidity, or funding pressure
- Competitive risks: market share loss or pricing pressure from identified competitors
- Regulatory or legal risks: specific laws, regulations, or litigation
- Macro or cyclical risks: only if quantifiably material to this company's financials
- Technology or disruption risks: new entrants, automation, business model displacement

THESIS-BREAKER TEST:
For each risk, ask: "If this risk fully crystallizes, does the investment thesis hold?"
If the answer is "no", flag it explicitly as a potential thesis-breaker and explain why.

RULES:
- Do not write a disclaimer. Write an analysis.
- Do not restate the thesis in negative form ("if growth does not materialize, the stock could underperform").
- Do not list risks without explaining the transmission mechanism.
- Do not write generic sector risks without company-specific quantification.
- Do not repeat risks that were already characterized as caveats in the highlight sections —
  this section should add new material, not summarize what was already said.
- Rank by materiality, not by type.
- 5 to 7 risks is the right number for a rigorous initiation. Fewer is too thin; more becomes a disclaimer.

STYLE:
- Institutional prose. You may use numbered risks for clarity and ranking signal.
- Write in the analytical third person or first-person plural as appropriate.
- Precise language: "a 100bps increase in labor costs could reduce EBITDA margins by ~40bps"
  is better than "labor cost inflation could weigh on margins."

LENGTH TARGET: 500 to 750 words

MASTER MEMO:
{memo}
"""
    system = (
        "You write risk sections for top-tier sell-side initiation reports. "
        "You are specific, ranked, and mechanistic. You never write a generic disclaimer "
        "and you never confuse 'naming a risk' with 'analyzing a risk'. "
        "A risk without a transmission mechanism is not an analysis."
    )
    return _chat(client, system, prompt, model=model, temperature=1)


# ---------------------------------------------------------------------------
# STEP 6 — FINAL EDITORIAL PASS
# ---------------------------------------------------------------------------

def final_edit_report(
    client: OpenAI,
    company_short_name: str,
    report_draft: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """
    Final editorial pass on the complete assembled draft.

    Role in the pipeline:
    - This is the last step before the report goes to Word templating.
    - The editor's job is to enforce cross-section consistency, eliminate redundancy,
      sharpen individual sentences, and ensure the report reads as one coherent voice.
    - The editor must NOT add new facts, change the thesis, or materially alter any section's argument.

    Prompt philosophy:
    - This is an editing pass, not a rewriting pass.
    - Specific targets: repeated phrases, inconsistent tone, AI-ish hedging language,
      promotional adjectives, and logical non-sequiturs between sections.
    - The final report should sound like it was written by one very good analyst, not
      assembled from six independently written sections.
    """
    prompt = f"""
You are the managing editor of the institutional research department reviewing a sell-side
initiation of coverage report on {company_short_name} before it is published.

Your task is a disciplined editorial pass. You are improving the draft — not rewriting it.
The analyst has done the thinking. Your job is to make the thinking cleaner, sharper,
and more consistent in voice across every section.

WHAT TO FIX:

1. CROSS-SECTION REDUNDANCY:
   Find and eliminate any observation that appears in substantially the same form
   in more than one section. Each section has a job — if an idea belongs in the
   company overview, cut it from the highlight section. If a risk was discussed as
   a caveat in a highlight, do not repeat it at the same level in the risks section.

2. TONE INCONSISTENCY:
   The report should sound like one voice: measured, analytical, institutional.
   Flag and fix: sections that are notably more promotional or more cautious than
   the overall tone. Normalize the first-person plural (we believe / we note / we estimate)
   consistently where it is used.

3. PROMOTIONAL LANGUAGE AND AI ARTIFACTS:
   Remove: "game-changing", "transformative", "unprecedented", "world-class", "best-in-class",
   "robust", "dynamic", "seamlessly", "holistic", "synergistic", "innovative", "cutting-edge".
   Remove: any sentence that reads like it was generated from a press release.
   Remove: any sentence that hedges analytically sound observations with unnecessary qualification.

4. SENTENCE PRECISION:
   Tighten: any sentence longer than 40 words that can be split without losing meaning.
   Remove: throat-clearing opener sentences ("It is worth noting that...", "As mentioned above...",
   "In summary...", "At the end of the day...").
   Tighten: any paragraph whose last sentence restates its first sentence.

5. LOGICAL FLOW:
   Check that each section ends in a way that naturally leads into the next.
   The report should have a narrative logic: thesis → highlights → context → deep-dives
   → financials → risks. If any section breaks the flow, adjust the transition.

6. FACTUAL CONSISTENCY:
   Check that specific numbers (revenue, margins, contract values, dates) are cited
   consistently across sections. If a number appears as "~$125M" in one section and
   "$125 million" in another, standardize it.

WHAT NOT TO CHANGE:
- Do not add new facts or analytical observations.
- Do not change the investment thesis unless the draft is internally inconsistent.
- Do not materially restructure any section — edit within the existing structure.
- Do not alter the five highlights — those were approved by the analyst.
- Do not soften the risks or strengthen the highlights based on your own view.

SELF-CHECK BEFORE OUTPUTTING:
After completing the edit, re-read the thesis and the first paragraph of each section.
Do they form a coherent, non-redundant analytical narrative? If not, continue editing.

OUTPUT:
Return only the fully edited report text. No commentary, no change log.

DRAFT REPORT:
{report_draft}
"""
    system = (
        "You are the head of editorial quality for an institutional equity research department. "
        "You have edited hundreds of initiation reports. You are meticulous, fast, and allergic "
        "to redundancy, promotional language, and analytical vagueness. "
        "You improve without rewriting. You sharpen without changing the argument."
    )
    return _chat(client, system, prompt, model=model, temperature=1)


# ---------------------------------------------------------------------------
# FORMATTING HELPERS
# ---------------------------------------------------------------------------

def format_highlights_for_prompt(highlights: Sequence[Dict[str, str]]) -> str:
    lines = []
    for idx, h in enumerate(highlights, 1):
        title = h.get("title", "").strip()
        highlight = h.get("highlight", "").strip()
        implication = h.get("implication", "").strip()
        angle = h.get("section_angle", "").strip()
        conviction = h.get("conviction_basis", "").strip()
        lines.append(
            f"{idx}. {title}\n"
            f"   Highlight   : {highlight}\n"
            f"   Implication : {implication}\n"
            f"   Angle       : {angle}\n"
            f"   Evidence    : {conviction}"
        )
    return "\n\n".join(lines)


# ---------------------------------------------------------------------------
# REPORT ASSEMBLY
# ---------------------------------------------------------------------------

def assemble_report_markdown(
    title: str,
    investment_thesis: Sequence[str],
    highlights: Sequence[Dict[str, str]],
    company_overview: str,
    highlight_sections: Sequence[Tuple[Dict[str, str], str]],
    financials: str,
    risks: str,
) -> str:
    """
    Assemble a clean Markdown report from all generated sections.
    This is intentionally simple so the Word templating phase can reuse the
    same structured content without re-parsing.
    """
    lines: List[str] = []
    lines.append(f"# {title}".strip())
    lines.append("")

    lines.append("## Investment Thesis")
    lines.append("")
    for p in investment_thesis:
        lines.append(p.strip())
        lines.append("")

    lines.append("## 5 Key Highlights")
    lines.append("")
    for idx, h in enumerate(highlights, 1):
        lines.append(f"**{idx}. {h.get('title', '').strip()}**")
        lines.append(f"{h.get('highlight', '').strip()}")
        if h.get("implication", "").strip():
            lines.append(f"*Implication: {h.get('implication', '').strip()}*")
        lines.append("")

    lines.append("## Company Overview")
    lines.append("")
    lines.append(company_overview.strip())
    lines.append("")

    for idx, (highlight, body) in enumerate(highlight_sections, 1):
        section_title = highlight.get("title", f"Highlight {idx}").strip()
        lines.append(f"## {idx}. {section_title}")
        lines.append("")
        lines.append(body.strip())
        lines.append("")

    lines.append("## Financials")
    lines.append("")
    lines.append(financials.strip())
    lines.append("")

    lines.append("## Risks")
    lines.append("")
    lines.append(risks.strip())
    lines.append("")

    return "\n".join(lines).strip() + "\n"


# ---------------------------------------------------------------------------
# I/O HELPERS
# ---------------------------------------------------------------------------

def save_text(path: str | Path, text: str) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def save_json(path: str | Path, data: Any) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
